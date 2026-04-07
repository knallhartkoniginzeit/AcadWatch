from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json
from datetime import date, datetime
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.neighbors import KNeighborsClassifier
from sklearn.preprocessing import StandardScaler

app = FastAPI(title="AcadWatch API v2")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"], allow_credentials=True)

EXCEL_PATH = os.environ.get("EXCEL_PATH", "acadwatch_data.xlsx")


SHEET_USERS         = "Users"
SHEET_STUDENTS      = "Students"
SHEET_SUBJECT_MARKS = "Subject_Marks"
SHEET_DAILY         = "Daily_Log"
SHEET_RECS          = "Recommendations"
SHEET_DASHBOARD     = "Dashboard_Summary"

SUBJECTS_LIST = ["TLA", "ML", "OS", "DE", "DPCO", "ML Lab", "OS Lab", "PDL", "SS", "CSBL"]

USER_COLS = ["username", "password", "role", "assigned_subjects", "student_id"]
SUBJECT_MARKS_COLS = ["student_id", "subject_name", "cbt1", "cat1", "assignment", "cbt2", "cat2", "final_mark", "updated_by"]

STUDENT_COLS = [
    "student_id","name","section","email",
    "internal_marks","previous_gpa","assignment_scores",
    "attendance_percentage","study_hours_per_day",
    "sleep_duration","mobile_usage_time","stress_level","interest_in_subject",
    "risk_label","risk_score","risk_probability",
    "last_updated","updated_by"
]
DAILY_COLS = [
    "date","student_id","name",
    "attendance_percentage","study_hours_per_day","sleep_duration",
    "mobile_usage_time","stress_level","internal_marks",
    "mood",
    "risk_label","risk_score","changed_from","updated_by"
]
REC_COLS = [
    "date","student_id","name","category","recommendation","status","updated_by"
]

# ─── Excel Initialiser ────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1E3A5F")
DANGER_FILL   = PatternFill("solid", start_color="FFE8EC")
WARNING_FILL  = PatternFill("solid", start_color="FFF4E0")
SUCCESS_FILL  = PatternFill("solid", start_color="E8FFF5")
ALT_FILL      = PatternFill("solid", start_color="F5F7FA")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT     = Font(name="Arial", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin", color="D0D7E2"),
    right=Side(style="thin", color="D0D7E2"),
    top=Side(style="thin", color="D0D7E2"),
    bottom=Side(style="thin", color="D0D7E2")
)

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data_row(ws, row_idx, risk_label=None):
    fill = BODY_FONT
    bg = ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
    if risk_label == "At Risk":
        bg = DANGER_FILL
    for cell in ws[row_idx]:
        cell.font = BODY_FONT
        cell.fill = bg
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER

def init_excel():
    if os.path.exists(EXCEL_PATH):
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Users sheet
    ws_u = wb.create_sheet(SHEET_USERS)
    ws_u.append(USER_COLS)
    style_header_row(ws_u)

    # Subject Marks sheet
    ws_sm = wb.create_sheet(SHEET_SUBJECT_MARKS)
    ws_sm.append(SUBJECT_MARKS_COLS)
    style_header_row(ws_sm)

    # Students sheet
    ws_s = wb.create_sheet(SHEET_STUDENTS)
    ws_s.append(STUDENT_COLS)
    style_header_row(ws_s)
    widths = [12,20,10,24,14,12,16,16,16,14,14,12,16,10,12,16,18,14]
    for i, w in enumerate(widths, 1):
        ws_s.column_dimensions[get_column_letter(i)].width = w
    ws_s.row_dimensions[1].height = 30
    ws_s.freeze_panes = "A2"

    # Daily log sheet
    ws_d = wb.create_sheet(SHEET_DAILY)
    ws_d.append(DAILY_COLS)
    style_header_row(ws_d)
    for i, w in enumerate([12,12,20,16,16,14,14,12,14,10,12,14,14], 1):
        ws_d.column_dimensions[get_column_letter(i)].width = w
    ws_d.row_dimensions[1].height = 30
    ws_d.freeze_panes = "A2"

    # Recommendations sheet
    ws_r = wb.create_sheet(SHEET_RECS)
    ws_r.append(REC_COLS)
    style_header_row(ws_r)
    for i, w in enumerate([12,12,20,16,60,14,14], 1):
        ws_r.column_dimensions[get_column_letter(i)].width = w
    ws_r.freeze_panes = "A2"

    # Dashboard summary sheet
    ws_dash = wb.create_sheet(SHEET_DASHBOARD)
    ws_dash.append(["Metric","Value","Date"])
    style_header_row(ws_dash)
    ws_dash.column_dimensions["A"].width = 30
    ws_dash.column_dimensions["B"].width = 15
    ws_dash.column_dimensions["C"].width = 15

    wb.save(EXCEL_PATH)

init_excel()

# ─── Ensure Demo Credentials Exist ───────────────────────────────────────────
DEMO_USERS = [
    {"username": "admin@school.edu",   "password": "admin123",   "role": "Admin",   "assigned_subjects": "", "student_id": ""},
]

# Add Faculty for all subjects
for sub in SUBJECTS_LIST:
    uname = f"{sub.replace(' ', '')}@school.edu"
    pwd = sub.replace(' ', '')
    DEMO_USERS.append({"username": uname, "password": pwd, "role": "Faculty", "assigned_subjects": sub, "student_id": ""})

# Add all 68 students
for i in range(1, 69):
    sid = f"student{i:02d}"
    uname = f"{sid}@school.edu"
    DEMO_USERS.append({"username": uname, "password": sid, "role": "Student", "assigned_subjects": "", "student_id": sid})

def ensure_default_users():
    """On every startup, make sure all demo users exist, have correct roles and student_ids, and are unique."""
    print(f"[Auth] Synchronizing user database at {EXCEL_PATH}...")
    try:
        users_df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_USERS, dtype={"username": str, "password": str, "student_id": str})
    except Exception as e:
        print(f"[Auth] Could not read Users sheet ({e}), creating fresh data.")
        users_df = pd.DataFrame(columns=USER_COLS)

    # Normalize existing data
    if not users_df.empty:
        # Clean current entries: strip all, lowercase username
        for col in USER_COLS:
            users_df[col] = users_df[col].astype(str).replace("nan", "").str.strip()
        users_df["username_lower"] = users_df["username"].str.lower()
    else:
        users_df["username_lower"] = []

    changed = False
    
    for demo in DEMO_USERS:
        d_uname_lower = str(demo["username"]).strip().lower()
        mask = users_df["username_lower"] == d_uname_lower
        
        if not mask.any():
            print(f"[Auth] Adding missing demo account: {demo['username']}")
            users_df = pd.concat([users_df, pd.DataFrame([demo])], ignore_index=True)
            users_df["username_lower"] = users_df["username"].astype(str).str.lower()
            changed = True
        else:
            # User exists - enforce correct Role, StudentID, and Password
            idx = users_df[mask].index[0]
            if users_df.at[idx, "password"] != str(demo["password"]).strip():
                users_df.at[idx, "password"] = str(demo["password"]).strip()
                changed = True
            if users_df.at[idx, "role"] != demo["role"]:
                users_df.at[idx, "role"] = demo["role"]
                changed = True
            if users_df.at[idx, "student_id"] != str(demo["student_id"]).strip():
                users_df.at[idx, "student_id"] = str(demo["student_id"]).strip()
                changed = True

    # Critical Step: Remove any duplicates (case-insensitive username)
    if users_df["username"].str.lower().duplicated().any():
        print(f"[Auth] Cleanup: Removing {users_df['username'].str.lower().duplicated().sum()} duplicate entries.")
        users_df = users_df.drop_duplicates(subset="username_lower", keep="last")
        changed = True

    if changed:
        print("[Auth] Saving 100% synchronized credentials to Excel...")
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if SHEET_USERS in wb.sheetnames:
            del wb[SHEET_USERS]
        ws_u = wb.create_sheet(SHEET_USERS)
        ws_u.append(USER_COLS)
        style_header_row(ws_u)
        
        for _, row in users_df.iterrows():
            ws_u.append([str(row.get(c, "")).strip() for c in USER_COLS])
            style_data_row(ws_u, ws_u.max_row)
            
        wb.save(EXCEL_PATH)
        print(f"[Auth] Synchronization complete. {len(users_df)} clean accounts verified.")
    else:
        print("[Auth] Global user synchronization verified: 100% match.")
    
    del users_df

ensure_default_users()


def compute_risk_components(d: dict):
    return {
        "Academic Risk": (max(0, 100 - d["internal_marks"]), 0.30),
        "Attendance Risk": (max(0, 100 - d["attendance_percentage"]), 0.25),
        "Study Hours Risk": (0 if d["study_hours_per_day"] >= 5 else (5 - d["study_hours_per_day"]) * 20, 0.15),
        "Sleep Risk": (0 if d["sleep_duration"] >= 7 else (7 - d["sleep_duration"]) * 15, 0.10),
        "Mobile Usage Risk": (0 if d["mobile_usage_time"] <= 3 else (d["mobile_usage_time"] - 3) * 10, 0.10),
        "Stress Risk": (d["stress_level"] * 10, 0.10),
    }

def compute_risk(d: dict, ml_prob: float = None):
    comps = compute_risk_components(d)
    rule_score = sum(val * weight for val, weight in comps.values())
    
    # Combine with ML logic if available
    if ml_prob is not None:
        final_score = (rule_score + (ml_prob * 100)) / 2.0
    else:
        final_score = rule_score
        
    final_score = round(final_score, 1)
    
    if final_score <= 10: label = "No Risk"
    elif final_score <= 25: label = "Below Risk"
    elif final_score <= 40: label = "Good Performance"
    elif final_score <= 60: label = "Can Improve"
    elif final_score <= 70: label = "Need Assistance"
    elif final_score <= 80: label = "At Risk"
    elif final_score <= 90: label = "High Risk"
    else: label = "Extremely High Risk"
    
    prob = round(final_score / 100.0, 3)
    return label, final_score, prob

def get_risk_factors(d):
    comps = compute_risk_components(d)
    # Calculate absolute contributions
    contributions = []
    for name, (val, weight) in comps.items():
        contrib = val * weight
        if contrib > 0:
            sev = "high" if val >= 60 else ("medium" if val >= 30 else "low")
            raw_val = d.get(name.lower().split(" ")[0] + "_percentage", d.get("internal_marks", 0)) if name == "Academic Risk" else val
            contributions.append({"factor": name, "value": f"Risk Score: {round(contrib,1)}", "severity": sev, "contrib": contrib})
            
    contributions.sort(key=lambda x: x["contrib"], reverse=True)
    
    # Return top 3 factors
    for c in contributions:
        del c["contrib"]
    return contributions[:3]

def build_recommendations(d, risk_factors):
    recs = []
    names = [r["factor"] for r in risk_factors]
    if "Attendance" in names:
        recs.append({"category":"Attendance","icon":"📅","title":"Improve Attendance",
            "actions":["Target minimum 85% attendance","Set class reminders","Buddy system with a peer","Discuss barriers with advisor"]})
    if "Internal Marks" in names or "Assignment Scores" in names:
        recs.append({"category":"Academics","icon":"📚","title":"Strengthen Academics",
            "actions":["30-min daily revision on weak topics","Practice previous year papers","Use Anki for spaced repetition","Join a study group"]})
    if "Mobile Usage" in names:
        recs.append({"category":"Focus","icon":"📵","title":"Reduce Screen Time",
            "actions":["Use Forest/Freedom during study hours","Keep phone in another room","Limit entertainment to 2 hrs/day","Enable Do Not Disturb"]})
    if "Sleep" in names:
        recs.append({"category":"Wellbeing","icon":"😴","title":"Better Sleep",
            "actions":["Target 7-8 hrs of sleep nightly","Consistent sleep/wake times","No screens 30 min before bed","20-min power nap for afternoon focus"]})
    if "Stress Level" in names:
        recs.append({"category":"Mental Health","icon":"🧘","title":"Manage Stress",
            "actions":["10-min daily mindfulness","Break tasks into milestones","Talk to a counselor","30-min daily walk"]})
    if "Study Hours" in names:
        recs.append({"category":"Study Habits","icon":"⏱️","title":"Build Study Routine",
            "actions":["3+ focused study hours daily","Pomodoro: 25 min on / 5 min break","Weekly timetable","Review notes within 24 hrs of class"]})
    if not recs:
        recs.append({"category":"Maintain","icon":"🌟","title":"Keep It Up!",
            "actions":["Maintain current habits","Help peers","Explore advanced topics","Set higher goals"]})
    return recs
# ─── ML Training Data (68 students) ─────────────────────────────────────────
# Columns: internal_marks, previous_gpa, assignment_scores, attendance_percentage,
#          study_hours_per_day, sleep_duration, mobile_usage_time, stress_level, interest_in_subject
ML_TRAIN_RAW = [
    [78,8.1,82,88,5,7,3,3,8],[45,5.2,50,55,1.5,4,9,9,3],[60,6.5,65,70,3,6,5,6,6],
    [82,8.8,85,92,6,7,2,2,9],[39,4.8,42,50,1,4,10,10,2],[67,7.0,70,75,4,6,4,5,7],
    [90,9.2,88,95,7,8,2,2,9],[52,6.0,58,68,3,5,6,7,5],[48,5.5,52,60,2,5,7,8,4],
    [73,7.8,76,85,5,7,3,4,8],[66,7.2,68,78,4,6,4,5,7],[35,4.5,40,48,1,3,11,9,2],
    [81,8.5,84,90,6,7,2,3,9],[59,6.3,62,72,3,6,5,6,6],[44,5.0,48,58,2,5,8,8,3],
    [77,8.0,80,88,5,7,3,4,8],[92,9.5,90,96,7,8,2,2,10],[50,6.0,55,65,2.5,5,6,7,5],
    [68,7.3,70,80,4,6,4,5,7],[38,4.7,42,52,1,4,9,9,3],[85,8.9,87,93,6,7,2,3,9],
    [62,6.8,64,74,3.5,6,5,6,6],[47,5.4,50,60,2,5,7,8,4],[79,8.2,82,89,5,7,3,3,8],
    [34,4.3,38,45,1,3,10,10,2],[69,7.4,72,82,4,6,4,5,7],[88,9.0,90,94,7,8,2,2,9],
    [55,6.2,58,68,3,5,6,7,5],[49,5.6,52,62,2,5,7,8,4],[74,7.9,78,86,5,7,3,4,8],
    [65,7.1,67,76,4,6,4,5,7],[36,4.6,40,50,1,4,10,9,3],[83,8.7,85,91,6,7,2,3,9],
    [58,6.1,60,70,3,6,5,6,6],[43,5.1,46,58,2,5,8,8,3],[76,8.0,79,88,5,7,3,4,8],
    [91,9.3,89,96,7,8,2,2,10],[51,6.0,54,66,2.5,5,6,7,5],[67,7.2,70,80,4,6,4,5,7],
    [37,4.7,41,52,1,4,9,9,3],[84,8.8,86,92,6,7,2,3,9],[61,6.7,63,74,3.5,6,5,6,6],
    [46,5.3,49,60,2,5,7,8,4],[78,8.1,81,89,5,7,3,3,8],[33,4.2,37,45,1,3,10,10,2],
    [70,7.5,73,82,4,6,4,5,7],[89,9.1,91,95,7,8,2,2,9],[56,6.3,59,69,3,5,6,7,5],
    [48,5.5,51,62,2,5,7,8,4],[75,7.9,78,87,5,7,3,4,8],[66,7.1,68,77,4,6,4,5,7],
    [35,4.5,39,50,1,4,10,9,3],[82,8.6,84,91,6,7,2,3,9],[57,6.2,60,71,3,6,5,6,6],
    [42,5.0,45,58,2,5,8,8,3],[77,8.0,80,88,5,7,3,4,8],[93,9.6,91,97,7,8,2,2,10],
    [52,6.1,55,67,2.5,5,6,7,5],[68,7.3,71,81,4,6,4,5,7],[38,4.6,42,53,1,4,9,9,3],
    [86,8.9,88,94,6,7,2,3,9],[63,6.9,65,75,3.5,6,5,6,6],[47,5.4,50,61,2,5,7,8,4],
    [80,8.3,83,90,5,7,3,3,8],[34,4.3,38,46,1,3,10,10,2],[71,7.6,74,83,4,6,4,5,7],
    [90,9.2,92,96,7,8,2,2,10],[54,6.2,57,68,3,5,6,7,5],
]

ML_STUDENT_NAMES = [
    "Student1","Student2","Student3","Student4","Student5","Student6","Student7","Student8",
    "Student9","Student10","Student11","Student12","Student13","Student14","Student15",
    "Student16","Student17","Student18","Student19","Student20","Student21","Student22",
    "Student23","Student24","Student25","Student26","Student27","Student28","Student29",
    "Student30","Student31","Student32","Student33","Student34","Student35","Student36",
    "Student37","Student38","Student39","Student40","Student41","Student42","Student43",
    "Student44","Student45","Student46","Student47","Student48","Student49","Student50",
    "Student51","Student52","Student53","Student54","Student55","Student56","Student57",
    "Student58","Student59","Student60","Student61","Student62","Student63","Student64",
    "Student65","Student66","Student67","Student68",
]


def _raw_to_dict(row):
    keys = ["internal_marks","previous_gpa","assignment_scores","attendance_percentage",
            "study_hours_per_day","sleep_duration","mobile_usage_time","stress_level","interest_in_subject"]
    return dict(zip(keys, row))


def _make_feature_vector(d: dict):
    return [
        d["internal_marks"], d["previous_gpa"], d["assignment_scores"],
        d["attendance_percentage"], d["study_hours_per_day"], d["sleep_duration"],
        d["mobile_usage_time"], d["stress_level"], d["interest_in_subject"],
    ]


# Build training labels using the rule-based engine, then train models
_X_train = np.array(ML_TRAIN_RAW, dtype=float)
_y_train = np.array(
    [1 if compute_risk(_raw_to_dict(r))[1] >= 61 else 0 for r in ML_TRAIN_RAW]
)

_scaler = StandardScaler()
_X_scaled = _scaler.fit_transform(_X_train)

rf_model  = RandomForestClassifier(n_estimators=200, random_state=42)
rf_model.fit(_X_train, _y_train)

lr_model  = LogisticRegression(max_iter=2000, random_state=42)
lr_model.fit(_X_scaled, _y_train)

knn_model = KNeighborsClassifier(n_neighbors=5)
knn_model.fit(_X_scaled, _y_train)

print(f"[ML] Models trained on {len(_X_train)} students | "
      f"At Risk: {_y_train.sum()} | Pass: {(1-_y_train).sum()}")


def compute_risk_ml(d: dict):
    """Returns (rf_label, lr_label, knn_label, ensemble_label, agreement_pct)."""
    fv = np.array([_make_feature_vector(d)], dtype=float)
    fv_s = _scaler.transform(fv)

    rf_pred  = int(rf_model.predict(fv)[0])
    lr_pred  = int(lr_model.predict(fv_s)[0])
    knn_pred = int(knn_model.predict(fv_s)[0])

    rf_label  = "At Risk" if rf_pred  else "Pass"
    lr_label  = "At Risk" if lr_pred  else "Pass"
    knn_label = "At Risk" if knn_pred else "Pass"

    votes_at_risk = rf_pred + lr_pred + knn_pred
    ensemble_label = "At Risk" if votes_at_risk >= 2 else "Pass"
    agreement_pct  = round(max(votes_at_risk, 3 - votes_at_risk) / 3 * 100, 1)

    # Probabilities
    rf_prob  = round(float(rf_model.predict_proba(fv)[0][1]),  3)
    lr_prob  = round(float(lr_model.predict_proba(fv_s)[0][1]), 3)
    knn_prob = round(float(knn_model.predict_proba(fv_s)[0][1]), 3)

    return rf_label, lr_label, knn_label, ensemble_label, agreement_pct, rf_prob, lr_prob, knn_prob


# ─── Schemas ─────────────────────────────────────────────────────────────────
class StudentInput(BaseModel):
    student_id: str
    name: str
    section: Optional[str] = ""
    email: Optional[str] = ""
    internal_marks: float
    previous_gpa: float
    assignment_scores: float
    attendance_percentage: float
    study_hours_per_day: float
    sleep_duration: float
    mobile_usage_time: float
    stress_level: int
    interest_in_subject: int
    updated_by: Optional[str] = "teacher"

class RecommendationUpdate(BaseModel):
    student_id: str
    category: str
    status: str
    updated_by: Optional[str] = "teacher"

class LoginRequest(BaseModel):
    role: str
    username: str
    password: str

class SubjectMarkInput(BaseModel):
    student_id: str
    subject_name: str
    cbt1: float
    cat1: float
    assignment: float
    cbt2: float
    cat2: float
    updated_by: Optional[str] = "faculty"

class DailyHabitInput(BaseModel):
    student_id: str
    sleep_hours: float   # 4–12 hours
    mood: int            # 1–5 scale

# ─── Helpers ──────────────────────────────────────────────────────────────────
def read_students() -> pd.DataFrame:
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_STUDENTS, dtype={"student_id":str})
        return df
    except:
        return pd.DataFrame(columns=STUDENT_COLS)

def read_daily() -> pd.DataFrame:
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_DAILY, dtype={"student_id":str})
        df["date"] = pd.to_datetime(df["date"]).dt.date
        return df
    except:
        return pd.DataFrame(columns=DAILY_COLS)

def read_recs() -> pd.DataFrame:
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_RECS, dtype={"student_id":str})
        return df
    except:
        return pd.DataFrame(columns=REC_COLS)

def save_students(df: pd.DataFrame):
    """Safely saves the Students sheet with retry logic for file locks."""
    import time
    for redo in range(3):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            if SHEET_STUDENTS in wb.sheetnames:
                del wb[SHEET_STUDENTS]
            ws = wb.create_sheet(SHEET_STUDENTS, 0)
            ws.append(STUDENT_COLS)
            style_header_row(ws)
            for _, row in df.iterrows():
                vals = [row.get(c,"") for c in STUDENT_COLS]
                ws.append(vals)
                risk = row.get("risk_label","")
                style_data_row(ws, ws.max_row, risk)
            widths = [12,20,10,24,14,12,16,16,16,14,14,12,16,10,12,16,18,14]
            for i,w in enumerate(widths,1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 30
            wb.save(EXCEL_PATH)
            print(f"[Admin] Successfully saved {len(df)} students to Excel.")
            return
        except Exception as e:
            print(f"[Admin] Attempt {redo+1} failed to save students: {e}")
            if redo == 2: raise e
            time.sleep(0.5)

def append_daily_log(row_data: dict):
    """Safely appends a row to the Daily_Log sheet with retry logic."""
    import time
    for redo in range(3):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb[SHEET_DAILY]
            ws.append([row_data.get(c,"") for c in DAILY_COLS])
            style_data_row(ws, ws.max_row, row_data.get("risk_label",""))
            wb.save(EXCEL_PATH)
            return
        except Exception as e:
            print(f"[Log] Attempt {redo+1} failed to append daily log: {e}")
            if redo == 2: raise e
            time.sleep(0.5)

def append_recommendations(student_id, name, recs, updated_by):
    """Safely appends recommendations with retry logic."""
    import time
    today = str(date.today())
    for redo in range(3):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb[SHEET_RECS]
            for rec in recs:
                for action in rec["actions"]:
                    ws.append([today, student_id, name, rec["category"], action, "Pending", updated_by])
                    style_data_row(ws, ws.max_row)
            wb.save(EXCEL_PATH)
            return
        except Exception as e:
            print(f"[Admin] Attempt {redo+1} failed to save recs: {e}")
            if redo == 2: raise e
            time.sleep(0.5)

def update_dashboard_summary():
    """Safely updates dashboard summary with retry logic."""
    import time
    for redo in range(3):
        try:
            df = read_students()
            if df.empty: return
            wb = openpyxl.load_workbook(EXCEL_PATH)
            if SHEET_DASHBOARD in wb.sheetnames: del wb[SHEET_DASHBOARD]
            ws = wb.create_sheet(SHEET_DASHBOARD)
            ws.append(["Metric","Value","Date"])
            style_header_row(ws)
            today = str(date.today())
            t = len(df); ar = len(df[df["risk_label"]=="At Risk"]); p = len(df[df["risk_label"]=="Pass"])
            rows = [
                ("Total Students", t, today), ("At Risk", ar, today), ("Passing", p, today),
                ("Risk Rate (%)", round(ar/t*100,1) if t else 0, today),
                ("Avg Attendance %", round(df["attendance_percentage"].mean(),1) if t else 0, today),
                ("Avg Risk Score", round(df["risk_score"].mean(),1) if t else 0, today),
            ]
            for i,(m,v,d) in enumerate(rows,2):
                ws.cell(i,1,m); ws.cell(i,2,v); ws.cell(i,3,d)
                for c in ws[i]:
                    c.font = BODY_FONT; c.border = THIN_BORDER
                    c.fill = ALT_FILL if i%2==0 else PatternFill("solid", start_color="FFFFFF")
            ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 14; ws.column_dimensions["C"].width = 14
            wb.save(EXCEL_PATH)
            return
        except Exception as e:
            print(f"[Admin] Attempt {redo+1} failed dashboard summary update: {e}")
            if redo == 2: raise e
            time.sleep(0.5)

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.post("/login")
def login(req: LoginRequest):
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_USERS, dtype={"username":str})
    except:
        df = pd.DataFrame(columns=USER_COLS)
        
    user = df[(df["username"].astype(str).str.lower() == req.username.lower()) & (df["role"].astype(str).str.lower() == req.role.lower())]
    if user.empty:
        raise HTTPException(status_code=401, detail="Invalid credentials")
        
    user = user.iloc[0]
    if str(user["password"]) != req.password:
        raise HTTPException(status_code=401, detail="Invalid credentials")
        
    resp = {"status": "success", "role": str(user["role"]).lower(), "username": req.username}
    if resp["role"] == "faculty":
        subs = str(user["assigned_subjects"]).split(",") if pd.notna(user["assigned_subjects"]) else []
        resp["assigned_subjects"] = [s.strip() for s in subs if s.strip()]
    elif resp["role"] == "student":
        resp["student_id"] = str(user["student_id"])
        
    return resp

def read_subject_marks() -> pd.DataFrame:
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_SUBJECT_MARKS, dtype={"student_id":str})
        return df
    except:
        return pd.DataFrame(columns=SUBJECT_MARKS_COLS)

def save_subject_marks(df: pd.DataFrame):
    """Safely saves the Subject_Marks sheet with retry logic."""
    import time
    for redo in range(3):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            if SHEET_SUBJECT_MARKS in wb.sheetnames:
                del wb[SHEET_SUBJECT_MARKS]
            ws = wb.create_sheet(SHEET_SUBJECT_MARKS)
            ws.append(SUBJECT_MARKS_COLS)
            style_header_row(ws)
            for _, row in df.iterrows():
                ws.append([row.get(c,"") for c in SUBJECT_MARKS_COLS])
                style_data_row(ws, ws.max_row)
            wb.save(EXCEL_PATH)
            return
        except Exception as e:
            print(f"[Faculty] Attempt {redo+1} failed mark save: {e}")
            if redo == 2: raise e
            time.sleep(0.5)

@app.get("/subjects/marks/{student_id}")
def get_student_marks(student_id: str):
    df = read_subject_marks()
    marks = df[df["student_id"] == student_id].to_dict("records")
    existing_subs = {m["subject_name"]: m for m in marks}
    result = []
    for sub in SUBJECTS_LIST:
        if sub in existing_subs:
            result.append(existing_subs[sub])
        else:
            result.append({"student_id": student_id, "subject_name": sub, "cbt1":0, "cat1":0, "assignment":0, "cbt2":0, "cat2":0, "final_mark":0})
    return result

@app.post("/subjects/marks")
def upsert_subject_marks_api(data: SubjectMarkInput):
    print(f"[Faculty] Updating marks for {data.student_id}, subject: {data.subject_name}")
    try:
        df = read_subject_marks()
        # CBT1 (30), CAT1 (60), Assignment (40), CBT2 (20), CAT2 (40). Total 190.
        total_score = data.cbt1 + data.cat1 + data.assignment + data.cbt2 + data.cat2
        final_mark = round((total_score * 100) / 190.0, 1)
        final_mark = min(100.0, final_mark) # Cap at 100
        
        new_row = {
            "student_id": data.student_id, "subject_name": data.subject_name,
            "cbt1": data.cbt1, "cat1": data.cat1, "assignment": data.assignment,
            "cbt2": data.cbt2, "cat2": data.cat2, "final_mark": final_mark,
            "updated_by": data.updated_by
        }
        
        mask = (df["student_id"].astype(str) == str(data.student_id)) & (df["subject_name"] == data.subject_name)
        if not df[mask].empty:
            df.loc[mask, list(new_row.keys())] = list(new_row.values())
        else:
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            
        save_subject_marks(df)
        
        # Cross-update the main student spreadsheet to trigger overall risk re-evaluation
        s_df = read_students()
        id_mask = s_df["student_id"].astype(str).str.strip().str.lower() == str(data.student_id).strip().lower()
        
        if not s_df[id_mask].empty:
            # Get latest average performance across ALL subjects for this student
            all_marks = df[df["student_id"].astype(str) == str(data.student_id)]
            avg_internal_mark = round(all_marks["final_mark"].mean(), 1)
            
            # Map existing student data to StudentInput model safely
            s_data = s_df[id_mask].iloc[0].to_dict()
            s_data = {k: (v if pd.notna(v) else "") for k, v in s_data.items()}
            s_data["internal_marks"] = avg_internal_mark
            s_data["updated_by"] = data.updated_by
            
            # Use only matching fields for the model to prevent validation errors
            fields = StudentInput.__fields__ if hasattr(StudentInput, "__fields__") else StudentInput.model_fields
            in_data = {k: v for k, v in s_data.items() if k in fields}
            
            st_in = StudentInput(**in_data)
            upsert_student(st_in)
            print(f"[Faculty] Auto-recalculated risk for Student {data.student_id} based on new marks.")
            
        return new_row
    except Exception as e:
        print(f"[Faculty] CRITICAL ERROR in upsert_subject_marks_api: {e}")
        raise HTTPException(500, detail=f"Failed to update marks: {str(e)}")

# ── Dashboard ──
@app.get("/dashboard")
def dashboard():
    df    = read_students()
    daily = read_daily()
    recs  = read_recs()
    today = date.today()

    if df.empty:
        return {
            "total":0,"at_risk":0,"passing":0,"risk_rate":0,
            "avg_attendance":0,"avg_risk_score":0,
            "worsened_today":[],"today_at_risk":[],"pending_recs":0,
            "attendance_trend":[],"risk_trend":[],"students":[]
        }

    total   = len(df)
    at_risk = int((df["risk_label"]=="At Risk").sum())
    passing = int((df["risk_label"]=="Pass").sum())
    avg_att = round(float(df["attendance_percentage"].mean()),1)
    avg_rs  = round(float(df["risk_score"].mean()),1)

    # Students who worsened today (Pass → At Risk)
    worsened = []
    if not daily.empty:
        today_log = daily[daily["date"]==today]
        for _, row in today_log.iterrows():
            if str(row.get("changed_from","")) == "Pass" and row.get("risk_label") == "At Risk":
                worsened.append({
                    "student_id": row["student_id"],
                    "name": row["name"],
                    "risk_score": row.get("risk_score",0),
                    "attendance_percentage": row.get("attendance_percentage",0),
                })

    # Today's at-risk list
    today_at_risk = df[df["risk_label"]=="At Risk"][["student_id","name","risk_score","attendance_percentage","section"]].to_dict("records")

    # Pending recommendations
    pending_recs = 0
    if not recs.empty and "status" in recs.columns:
        pending_recs = int((recs["status"]=="Pending").sum())

    # Attendance trend (last 7 days from daily log)
    att_trend = []
    if not daily.empty:
        daily["date"] = pd.to_datetime(daily["date"]).dt.date
        last7 = sorted(daily["date"].unique())[-7:]
        for d in last7:
            day_data = daily[daily["date"]==d]
            avg = round(float(day_data["attendance_percentage"].mean()),1) if len(day_data) else None
            att_trend.append({"date": str(d), "avg_attendance": avg})

    # Risk score trend
    risk_trend = []
    if not daily.empty:
        last7 = sorted(daily["date"].unique())[-7:]
        for d in last7:
            day_data = daily[daily["date"]==d]
            avg = round(float(day_data["risk_score"].mean()),1) if len(day_data) else None
            at_r = int((day_data["risk_label"]=="At Risk").sum()) if len(day_data) else 0
            risk_trend.append({"date":str(d),"avg_risk_score":avg,"at_risk_count":at_r})

    students = df.replace({float("nan"):None}).to_dict("records")

    return {
        "total": total, "at_risk": at_risk, "passing": passing,
        "risk_rate": round(at_risk/total*100,1) if total else 0,
        "avg_attendance": avg_att, "avg_risk_score": avg_rs,
        "worsened_today": worsened,
        "today_at_risk": today_at_risk,
        "pending_recs": pending_recs,
        "attendance_trend": att_trend,
        "risk_trend": risk_trend,
        "students": students,
    }

# ── Students CRUD ──
@app.get("/students")
def list_students():
    df = read_students()
    return df.replace({float("nan"):None}).to_dict("records")

@app.get("/students/{student_id}")
def get_student(student_id: str):
    df = read_students()
    row = df[df["student_id"]==student_id]
    if row.empty:
        raise HTTPException(404,"Student not found")
    return row.replace({float("nan"):None}).iloc[0].to_dict()

def append_daily_log(row_data: dict):
    """Safely appends a row to the Daily_Log sheet with retry logic for file locks."""
    import time
    for redo in range(3):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            if SHEET_DAILY not in wb.sheetnames:
                ws = wb.create_sheet(SHEET_DAILY)
                ws.append(DAILY_COLS)
                style_header_row(ws)
            else:
                ws = wb[SHEET_DAILY]
            
            ws.append([row_data.get(c, "") for c in DAILY_COLS])
            style_data_row(ws, ws.max_row)
            wb.save(EXCEL_PATH)
            print(f"[Log] Successfully appended daily entry for {row_data.get('student_id')}")
            return
        except Exception as e:
            print(f"[Log] Attempt {redo+1} failed to save daily log: {e}")
            time.sleep(0.5)

@app.post("/students")
def upsert_student(data: StudentInput):
    print(f"[Admin] Upserting student: {data.student_id}")
    try:
        df = read_students()
        
        # First get ML predictions
        rf_label, lr_label, knn_label, ensemble_label, agreement_pct, rf_prob, lr_prob, knn_prob = compute_risk_ml(data.dict())
        
        # Compute combined final risk using ML probability (e.g. Random Forest prob)
        label, score, prob = compute_risk(data.dict(), ml_prob=rf_prob)
        
        risk_factors = get_risk_factors(data.dict())
        recommendations = build_recommendations(data.dict(), risk_factors)
        now = datetime.now().strftime("%Y-%m-%d %H:%M")

        prev_label = None
        if not df.empty and data.student_id in df["student_id"].values:
            prev_label = df.loc[df["student_id"]==data.student_id,"risk_label"].values[0]

        new_row = {
            "student_id": data.student_id, "name": data.name,
            "section": data.section, "email": data.email,
            "internal_marks": data.internal_marks, "previous_gpa": data.previous_gpa,
            "assignment_scores": data.assignment_scores,
            "attendance_percentage": data.attendance_percentage,
            "study_hours_per_day": data.study_hours_per_day,
            "sleep_duration": data.sleep_duration,
            "mobile_usage_time": data.mobile_usage_time,
            "stress_level": data.stress_level,
            "interest_in_subject": data.interest_in_subject,
            "risk_label": label, "risk_score": score, "risk_probability": prob,
            "last_updated": now, "updated_by": data.updated_by,
        }

        if not df.empty and data.student_id in df["student_id"].values:
            df.loc[df["student_id"]==data.student_id, list(new_row.keys())] = list(new_row.values())
        else:
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

        save_students(df)

        # Mirror entry in daily log for history
        append_daily_log({
            "date": str(date.today()), "student_id": data.student_id, "name": data.name,
            "attendance_percentage": data.attendance_percentage,
            "study_hours_per_day": data.study_hours_per_day,
            "sleep_duration": data.sleep_duration,
            "mobile_usage_time": data.mobile_usage_time,
            "stress_level": data.stress_level,
            "internal_marks": data.internal_marks,
            "risk_label": label, "risk_score": score,
            "changed_from": prev_label or "", "updated_by": data.updated_by,
        })

        if label == "At Risk":
            append_recommendations(data.student_id, data.name, recommendations, data.updated_by)

        update_dashboard_summary()

        return {
            "student_id": data.student_id, "name": data.name,
            "risk_label": label, "risk_score": score, "risk_probability": prob,
            "top_risk_factors": risk_factors, "recommendations": recommendations,
            "changed_from": prev_label,
            "rf_label": rf_label, "rf_probability": rf_prob
        }
    except Exception as e:
        print(f"[Admin] CRITICAL ERROR in upsert_student: {e}")
        raise HTTPException(500, detail=f"Data update failed: {str(e)}")

# ── Recommendations ──
@app.get("/recommendations")
def list_recommendations(student_id: Optional[str] = None):
    df = read_recs()
    if df.empty:
        return []
    if student_id:
        df = df[df["student_id"]==student_id]
        
    # Only show recommendations for students currently "At Risk"
    s_df = read_students()
    if not s_df.empty:
        at_risk_ids = s_df[s_df["risk_label"] == "At Risk"]["student_id"].tolist()
        df = df[df["student_id"].isin(at_risk_ids)]
        
    return df.replace({float("nan"):None}).to_dict("records")

@app.put("/recommendations/status")
def update_rec_status(data: RecommendationUpdate):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_RECS]
    headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    sid_col = headers.index("student_id")+1
    cat_col = headers.index("category")+1
    sta_col = headers.index("status")+1
    upd_col = headers.index("updated_by")+1
    for row in ws.iter_rows(min_row=2):
        if (str(row[sid_col-1].value)==data.student_id and
                str(row[cat_col-1].value)==data.category):
            row[sta_col-1].value  = data.status
            row[upd_col-1].value  = data.updated_by
    wb.save(EXCEL_PATH)
    return {"status":"updated"}

# ── Export ──
@app.get("/export/daily-report")
def export_daily():
    """Returns today's summary as JSON for the frontend to display."""
    daily = read_daily()
    today = date.today()
    if daily.empty:
        return {"date": str(today), "entries": []}
    today_data = daily[daily["date"]==today]
    return {"date": str(today), "entries": today_data.replace({float("nan"):None}).to_dict("records")}

@app.get("/export/excel-path")
def excel_path():
    return {"path": os.path.abspath(EXCEL_PATH)}

@app.post("/daily-log")
def add_daily_habit(data: DailyHabitInput):
    print(f"[Student] Logging habit for: {data.student_id}")
    try:
        df = read_students()
        # Robust ID lookup: trim, stringify, and lowercase both sides
        target_id = str(data.student_id).strip().lower()
        id_series = df["student_id"].astype(str).str.strip().str.lower()
        
        student_row = df[id_series == target_id]
        
        if student_row.empty:
            print(f"[Student] ERROR: ID '{target_id}' not found in students sheet. Count: {len(df)}")
            raise HTTPException(404, f"Student ID '{target_id}' not found. Please log in again.")
        
        s = student_row.iloc[0]
        name = str(s.get("name", "Student"))
        today = str(date.today())
        
        row_data = {
            "date": today, "student_id": data.student_id, "name": name,
            "attendance_percentage": s.get("attendance_percentage", 0),
            "study_hours_per_day": s.get("study_hours_per_day", 0),
            "sleep_duration": data.sleep_hours,
            "mobile_usage_time": s.get("mobile_usage_time", 0),
            "stress_level": s.get("stress_level", 0),
            "internal_marks": s.get("internal_marks", 0),
            "mood": data.mood,
            "risk_label": s.get("risk_label", ""),
            "risk_score": s.get("risk_score", 0),
            "changed_from": "", "updated_by": "student",
        }
        
        append_daily_log(row_data)
        return {"status": "logged", "date": today, "sleep_hours": data.sleep_hours, "mood": data.mood}
    except Exception as e:
        print(f"[Student] CRITICAL ERROR in add_daily_habit: {e}")
        raise HTTPException(500, detail=f"Logging failed: {str(e)}")

@app.get("/daily-log/{student_id}")
def get_daily_habit(student_id: str, days: int = 30):
    df = read_daily()
    if df.empty:
        return []
    df_s = df[df["student_id"] == student_id].copy()
    if df_s.empty:
        return []
    df_s = df_s.sort_values("date").tail(days)
    cols = [c for c in ["date","sleep_duration","mood","risk_label","risk_score"] if c in df_s.columns]
    result = df_s[cols].copy()
    result["date"] = result["date"].astype(str)
    return result.replace({float("nan"): None}).to_dict("records")

# ── Subject Stats (Admin Analytics) ──
@app.get("/subject-stats")
def get_subject_stats():
    marks_df = read_subject_marks()
    if marks_df.empty:
        return []
    result = []
    for sub in SUBJECTS_LIST:
        sub_df = marks_df[marks_df["subject_name"] == sub]
        if sub_df.empty:
            result.append({"subject": sub, "avg_mark": 0, "at_risk_count": 0, "pass_count": 0, "total": 0})
            continue
        avg = round(float(sub_df["final_mark"].mean()), 1)
        at_risk = int((sub_df["final_mark"] < 50).sum())
        passing = int((sub_df["final_mark"] >= 50).sum())
        result.append({"subject": sub, "avg_mark": avg, "at_risk_count": at_risk, "pass_count": passing, "total": len(sub_df)})
    return result

# ── Peer Groups (Community Feature) ──
@app.get("/peer-groups")
def get_peer_groups(subject: Optional[str] = None):
    students_df = read_students()
    marks_df = read_subject_marks()
    if students_df.empty:
        return []
    subjects = [subject] if subject else SUBJECTS_LIST
    all_groups = []
    for sub in subjects:
        sub_marks = marks_df[marks_df["subject_name"] == sub] if not marks_df.empty else pd.DataFrame()
        if sub_marks.empty:
            at_risk_list = students_df[students_df["risk_label"] == "At Risk"].replace({float("nan"): None}).to_dict("records")
            high_perf_list = students_df[students_df["risk_label"] == "Pass"].sort_values("risk_score", ascending=False).replace({float("nan"): None}).to_dict("records")
            for r in at_risk_list:  r["final_mark"] = r.get("risk_score", 0)
            for r in high_perf_list: r["final_mark"] = r.get("risk_score", 0)
        else:
            merged = sub_marks.merge(
                students_df[["student_id","name","section","risk_label","risk_score"]],
                on="student_id", how="left"
            ).replace({float("nan"): None})
            at_risk_list  = merged[merged["final_mark"] < 50].to_dict("records")
            high_perf_list = merged[merged["final_mark"] >= 70].sort_values("final_mark", ascending=False).to_dict("records")
        at_idx, hp_idx, group_idx = 0, 0, 0
        while at_idx < len(at_risk_list) or hp_idx < len(high_perf_list):
            learners = at_risk_list[at_idx:at_idx+3]
            mentors  = high_perf_list[hp_idx:hp_idx+2]
            if not learners and not mentors: break
            group_idx += 1
            all_groups.append({
                "group_id": f"{sub.replace(' ','_')}_G{group_idx}",
                "subject": sub, "group_number": group_idx,
                "learners": [{"student_id": m["student_id"], "name": m.get("name",""), "section": m.get("section",""), "final_mark": m.get("final_mark",0), "role": "Learner"} for m in learners],
                "mentors":  [{"student_id": m["student_id"], "name": m.get("name",""), "section": m.get("section",""), "final_mark": m.get("final_mark",0), "role": "Mentor"} for m in mentors],
                "members":  [{"student_id": m["student_id"], "name": m.get("name",""), "section": m.get("section",""), "final_mark": m.get("final_mark",0), "role": "Learner"} for m in learners] +
                            [{"student_id": m["student_id"], "name": m.get("name",""), "section": m.get("section",""), "final_mark": m.get("final_mark",0), "role": "Mentor"} for m in mentors],
            })
            at_idx += 3
            hp_idx += 2
    return all_groups

# ── Faculty: get students for assigned subjects ──
@app.get("/faculty/students")
def get_faculty_students(subjects: str = ""):
    students_df = read_students()
    marks_df    = read_subject_marks()
    if students_df.empty:
        return []
    sub_list = [s.strip() for s in subjects.split(",") if s.strip()]
    if not sub_list or marks_df.empty:
        return students_df.replace({float("nan"): None}).to_dict("records")
    relevant = marks_df[marks_df["subject_name"].isin(sub_list)]["student_id"].unique()
    filtered = students_df[students_df["student_id"].isin(relevant)]
    result = filtered.replace({float("nan"): None}).to_dict("records")
    for s in result:
        sid   = s["student_id"]
        sub_m = marks_df[(marks_df["student_id"] == sid) & (marks_df["subject_name"].isin(sub_list))]
        s["subject_marks"] = sub_m.replace({float("nan"): None}).to_dict("records")
    return result

# ── Seed demo data ──
@app.post("/seed")
def seed():
    demo = [
        StudentInput(student_id="CS21001",name="Ananya Sharma",section="A",email="ananya@college.edu",
            internal_marks=42,previous_gpa=5.2,assignment_scores=55,attendance_percentage=68,
            study_hours_per_day=1.5,sleep_duration=5.5,mobile_usage_time=8,stress_level=8,
            interest_in_subject=4,updated_by="teacher"),
        StudentInput(student_id="CS21002",name="Nikitha Reddy",section="A",email="nikitha@college.edu",
            internal_marks=78,previous_gpa=7.8,assignment_scores=85,attendance_percentage=91,
            study_hours_per_day=5,sleep_duration=7.5,mobile_usage_time=3,stress_level=4,
            interest_in_subject=8,updated_by="teacher"),
        StudentInput(student_id="CS21003",name="Siddeshwaran K",section="B",email="sidd@college.edu",
            internal_marks=55,previous_gpa=6.1,assignment_scores=60,attendance_percentage=79,
            study_hours_per_day=3,sleep_duration=6,mobile_usage_time=5,stress_level=6,
            interest_in_subject=6,updated_by="teacher"),
        StudentInput(student_id="CS21004",name="Priya Venkat",section="B",email="priya@college.edu",
            internal_marks=35,previous_gpa=4.5,assignment_scores=40,attendance_percentage=60,
            study_hours_per_day=1,sleep_duration=5,mobile_usage_time=10,stress_level=9,
            interest_in_subject=3,updated_by="teacher"),
        StudentInput(student_id="CS21005",name="Arjun Mehta",section="A",email="arjun@college.edu",
            internal_marks=88,previous_gpa=8.9,assignment_scores=92,attendance_percentage=96,
            study_hours_per_day=6,sleep_duration=8,mobile_usage_time=2,stress_level=3,
            interest_in_subject=9,updated_by="teacher"),
        StudentInput(student_id="CS21006",name="Divya Nair",section="B",email="divya@college.edu",
            internal_marks=48,previous_gpa=5.8,assignment_scores=52,attendance_percentage=72,
            study_hours_per_day=2,sleep_duration=6.5,mobile_usage_time=7,stress_level=7,
            interest_in_subject=5,updated_by="teacher"),
    ]
    results = []
    for s in demo:
        results.append(upsert_student(s))
    seed_users_and_marks(results)
    return {"seeded": len(results), "results": results}


# ── Bulk seed: all 68 ML-training students ──
def seed_users_and_marks(results):
    import random
    new_users = list(DEMO_USERS)
    
    # Remove the placeholder "student1@school.edu" from DEMO_USERS copy and re-add with first_id
    new_users = [u for u in new_users if u["username"] != "student01@school.edu"]
    
    if results:
        first_id = str(list(results)[0]["student_id"])
        new_users.append({
            "username": "student01@school.edu", "password": "student01",
            "role": "Student", "assigned_subjects": "", "student_id": first_id
        })
    # All students also accessible by student_id directly
    for r in list(results):
        sid_lower = str(r["student_id"]).lower()
        new_users.append({
            "username": f"{sid_lower}@school.edu", "password": sid_lower,
            "role": "Student", "assigned_subjects": "", "student_id": str(r["student_id"])
        })
    users_df = pd.DataFrame(new_users)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if SHEET_USERS in wb.sheetnames: del wb[SHEET_USERS]
    ws_u = wb.create_sheet(SHEET_USERS)
    ws_u.append(USER_COLS)
    style_header_row(ws_u)
    for _, row in users_df.iterrows():
        ws_u.append([row.get(c,"") for c in USER_COLS])
        style_data_row(ws_u, ws_u.max_row)
        
    # Also seed dummy marks for all students for all subjects
    if SHEET_SUBJECT_MARKS in wb.sheetnames: del wb[SHEET_SUBJECT_MARKS]
    ws_m = wb.create_sheet(SHEET_SUBJECT_MARKS)
    ws_m.append(SUBJECT_MARKS_COLS)
    style_header_row(ws_m)
    
    marks_rows = []
    for r in list(results):
        sid = r["student_id"]
        # Use internal_marks percent as base proxy for dummy marks
        base_pct = r.get("internal_marks", 60)
        for sub in SUBJECTS_LIST:
            # Add some controlled randomness around their base percent
            var = random.uniform(0.8, 1.2)
            cur_pct = min(100, max(0, base_pct * var))
            
            # Map score to the 190 total point scale
            cbt1 = round((30 * cur_pct) / 100)
            cat1 = round((60 * cur_pct) / 100)
            assign = round((40 * cur_pct) / 100)
            cbt2 = round((20 * cur_pct) / 100)
            cat2 = round((40 * cur_pct) / 100)
            
            total_190 = cbt1 + cat1 + assign + cbt2 + cat2
            final_mark = min(100.0, round((total_190 * 100) / 190.0, 1))
            
            marks_rows.append({
                "student_id": sid, "subject_name": sub, "cbt1": cbt1, "cat1": cat1, 
                "assignment": assign, "cbt2": cbt2, "cat2": cat2, "final_mark": final_mark, 
                "updated_by": "system"
            })
            
    for row_dict in marks_rows:
        ws_m.append([row_dict.get(c, "") for c in SUBJECT_MARKS_COLS])
        style_data_row(ws_m, ws_m.max_row)
        
    wb.save(EXCEL_PATH)

@app.post("/seed-bulk")
def seed_bulk():
    sections = ["A","B","C"]
    results = []
    for i, (row, name) in enumerate(zip(ML_TRAIN_RAW, ML_STUDENT_NAMES), start=1):
        d = _raw_to_dict(row)
        sid = f"student{i:02d}"
        sec = sections[(i-1) % 3]
        s = StudentInput(
            student_id=sid, name=name, section=sec,
            email=f"{name.lower().replace(' ','')}@college.edu",
            internal_marks=d["internal_marks"],
            previous_gpa=d["previous_gpa"],
            assignment_scores=d["assignment_scores"],
            attendance_percentage=d["attendance_percentage"],
            study_hours_per_day=d["study_hours_per_day"],
            sleep_duration=d["sleep_duration"],
            mobile_usage_time=d["mobile_usage_time"],
            stress_level=int(d["stress_level"]),
            interest_in_subject=int(d["interest_in_subject"]),
            updated_by="system",
        )
        results.append(upsert_student(s))
    seed_users_and_marks(results)
    return {"seeded": len(results), "at_risk": sum(1 for r in results if r["risk_label"]=="At Risk")}

# ── Frontend Catch-all ──
@app.get("/{full_path:path}")
async def serve_react(full_path: str):
    static_file = os.path.join("static", full_path)
    if os.path.isfile(static_file):
        return FileResponse(static_file)
    
    index_file = os.path.join("static", "index.html")
    if os.path.isfile(index_file):
        return FileResponse(index_file)
        
    return {"status": "API active. Frontend not built. Run 'npm run build' and put inside /static folder."}
